from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.shortcuts import redirect, render
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.utils import timezone

from core.mixins import GestionRequiredMixin, SuperAdminRequiredMixin, AdminRequiredMixin
from apps.gares.models import Gare
from apps.lignes.models import Ligne
from .models import Destination
from .forms import DestinationForm

# En-têtes attendus dans le fichier Excel d'import/export.
EXPORT_HEADERS = ['Gare (code)', 'Gare (nom)', 'Ligne', "Ville d'arrivée", 'Montant', 'Active']


class DestinationListView(GestionRequiredMixin, ListView):
    """Liste des destinations."""
    model = Destination
    template_name = 'destinations/destination_list.html'
    context_object_name = 'destinations'
    paginate_by = 15

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        # Filtrer par gare pour les chefs de gare et guichetiers
        if not user.has_global_access and user.gare:
            queryset = queryset.filter(gare=user.gare)

        # Recherche
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(
                Q(ville_arrivee__icontains=search) |
                Q(ligne__nom__icontains=search)
            )

        return queryset.select_related('ligne', 'gare').order_by('ligne__nom', 'ville_arrivee')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        context['total_count'] = context['paginator'].count
        return context


class DestinationCreateView(GestionRequiredMixin, CreateView):
    """Créer une nouvelle destination."""
    model = Destination
    form_class = DestinationForm
    template_name = 'destinations/destination_form.html'
    success_url = reverse_lazy('destinations:destination_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Destination créée avec succès.')
        return super().form_valid(form)


class DestinationUpdateView(GestionRequiredMixin, UpdateView):
    """Modifier une destination."""
    model = Destination
    form_class = DestinationForm
    template_name = 'destinations/destination_form.html'
    success_url = reverse_lazy('destinations:destination_list')

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        # Filtrer par gare pour les chefs de gare et guichetiers
        if not user.has_global_access and user.gare:
            queryset = queryset.filter(gare=user.gare)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Destination modifiée avec succès.')
        return super().form_valid(form)


class DestinationDeleteView(SuperAdminRequiredMixin, DeleteView):
    """
    Désactive une destination (soft-delete).
    La suppression physique est volontairement désactivée : une destination
    peut déjà être référencée par des voyages passés, la supprimer casserait
    cet historique. On la marque simplement inactive.
    """
    model = Destination
    template_name = 'destinations/destination_confirm_delete.html'
    success_url = reverse_lazy('destinations:destination_list')

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        # Filtrer par gare pour les chefs de gare et guichetiers
        if not user.has_global_access and user.gare:
            queryset = queryset.filter(gare=user.gare)

        return queryset

    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.active = False
        self.object.save(update_fields=['active'])
        messages.success(self.request, 'Destination désactivée avec succès.')
        return HttpResponseRedirect(success_url)


class DestinationExportView(AdminRequiredMixin, View):
    """
    Exporte les destinations (filtrées par la recherche en cours) en Excel.
    Le fichier généré peut être réédité (ex: changer la colonne "Gare (code)"
    pour dupliquer un jeu de destinations vers une autre gare) puis réimporté.
    """

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Destinations'
        ws.append(EXPORT_HEADERS)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3260', end_color='1E3260', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        if request.GET.get('modele') == '1':
            # Modèle vierge avec une ligne d'exemple.
            ws.append(['ADJ', 'Adjamé Nouvelle Gare', 'Abidjan - Conakry', 'Conakry', 45000, 'Oui'])
        else:
            queryset = Destination.objects.select_related('ligne', 'gare')
            search = request.GET.get('q')
            if search:
                queryset = queryset.filter(
                    Q(ville_arrivee__icontains=search) |
                    Q(ligne__nom__icontains=search)
                )
            queryset = queryset.order_by('gare__code', 'ligne__nom', 'ville_arrivee')
            for destination in queryset:
                ws.append([
                    destination.gare.code,
                    destination.gare.nom,
                    destination.ligne.nom,
                    destination.ville_arrivee,
                    float(destination.montant),
                    'Oui' if destination.active else 'Non',
                ])

        for i, width in enumerate([14, 26, 26, 20, 12, 10], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"destinations_{timezone.now():%Y%m%d_%H%M}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class DestinationImportView(AdminRequiredMixin, View):
    """
    Importe des destinations depuis un fichier Excel (même format que l'export).
    Upsert sur la clé (gare, ligne, ville_arrivee) : une combinaison déjà
    existante est mise à jour (montant/statut), une nouvelle combinaison
    (typiquement une gare différente) est créée.
    """
    template_name = 'destinations/destination_import.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        fichier = request.FILES.get('fichier')
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx).")
            return redirect('destinations:destination_import')

        try:
            wb = openpyxl.load_workbook(fichier, data_only=True)
        except Exception:
            messages.error(
                request,
                "Fichier illisible. Utilisez un fichier .xlsx généré par l'export de cette page."
            )
            return redirect('destinations:destination_import')

        ws = wb.active
        header_row = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
        try:
            idx_gare = header_row.index('Gare (code)')
            idx_ligne = header_row.index('Ligne')
            idx_ville = header_row.index("Ville d'arrivée")
            idx_montant = header_row.index('Montant')
            idx_active = header_row.index('Active')
        except ValueError:
            messages.error(
                request,
                "En-têtes de colonnes non reconnus. Utilisez le fichier exporté depuis cette page comme modèle."
            )
            return redirect('destinations:destination_import')

        nb_crees = 0
        nb_maj = 0
        erreurs = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None or str(v).strip() == '' for v in row):
                continue

            code_gare = str(row[idx_gare]).strip() if row[idx_gare] is not None else ''
            nom_ligne = str(row[idx_ligne]).strip() if row[idx_ligne] is not None else ''
            ville = str(row[idx_ville]).strip() if row[idx_ville] is not None else ''
            montant_brut = row[idx_montant]
            active_brut = str(row[idx_active]).strip().lower() if row[idx_active] is not None else 'oui'

            if not code_gare or not nom_ligne or not ville:
                erreurs.append(f"Ligne {row_num} : gare, ligne ou ville d'arrivée manquante.")
                continue

            gare = Gare.objects.filter(code__iexact=code_gare).first()
            if not gare:
                erreurs.append(f"Ligne {row_num} : gare '{code_gare}' introuvable.")
                continue

            ligne = Ligne.objects.filter(nom__iexact=nom_ligne, active=True).first()
            if not ligne:
                erreurs.append(f"Ligne {row_num} : ligne '{nom_ligne}' introuvable ou inactive.")
                continue

            try:
                montant = Decimal(str(montant_brut).replace(',', '.').replace(' ', ''))
                if montant < 0:
                    raise ValueError
            except (InvalidOperation, ValueError, TypeError):
                erreurs.append(f"Ligne {row_num} : montant '{montant_brut}' invalide.")
                continue

            active = active_brut in ('oui', 'yes', 'true', '1', 'actif', 'active')

            # Comparaison insensible à la casse pour éviter les doublons
            # (ex: "Kindia" et "KINDIA" doivent être vus comme la même destination).
            existante = Destination.objects.filter(
                gare=gare, ligne=ligne, ville_arrivee__iexact=ville
            ).first()
            if existante:
                existante.montant = montant
                existante.active = active
                existante.save(update_fields=['montant', 'active'])
                nb_maj += 1
            else:
                Destination.objects.create(
                    gare=gare, ligne=ligne, ville_arrivee=ville,
                    montant=montant, active=active,
                )
                nb_crees += 1

        if nb_crees or nb_maj:
            messages.success(
                request,
                f"Import terminé : {nb_crees} destination(s) créée(s), {nb_maj} mise(s) à jour."
            )
        if erreurs:
            apercu = " | ".join(erreurs[:15])
            suffixe = " ..." if len(erreurs) > 15 else ""
            messages.warning(request, f"{len(erreurs)} ligne(s) ignorée(s) : {apercu}{suffixe}")
        if not nb_crees and not nb_maj and not erreurs:
            messages.warning(request, "Le fichier ne contenait aucune ligne à importer.")

        return redirect('destinations:destination_list')
